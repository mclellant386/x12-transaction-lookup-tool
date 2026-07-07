import csv
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

project_folder = Path(__file__).parent
input_file = project_folder / "edi_transaction_report.csv"
output_file = project_folder / "edi_processing_report.xlsx"

total_transactions = 0
accepted_count = 0
rejected_count = 0
file_status = {}
rejection_summary = {}
rejection_detail_by_file = {}

with open(input_file, "r") as file:
    reader = csv.DictReader(file)

    for row in reader:
        total_transactions += 1
        file_name = row["file_name"]

        if file_name not in file_status:
            file_status[file_name] = True

        if row["status"] == "Accepted":
            accepted_count += 1

        elif row["status"] == "Rejected":
            rejected_count += 1
            file_status[file_name] = False

            rejection_key = (row["status_code"], row["status_message"])

            rejection_summary[rejection_key] = (
                rejection_summary.get(rejection_key, 0) + 1
            )

            if file_name not in rejection_detail_by_file:
                rejection_detail_by_file[file_name] = {}

            rejection_detail_by_file[file_name][rejection_key] = (
                rejection_detail_by_file[file_name].get(rejection_key, 0) + 1
    )

fully_accepted_files = []
files_with_rejections = []

for file_name, is_fully_accepted in file_status.items():
    if is_fully_accepted:
        fully_accepted_files.append(file_name)
    else:
        files_with_rejections.append(file_name)

acceptance_rate = accepted_count / total_transactions if total_transactions > 0 else 0

workbook = Workbook()
title_font = Font(size=16, bold=True)

section_font = Font(size=12, bold=True)

header_fill = PatternFill(
    fill_type="solid",
    start_color="D9EAF7",
    end_color="D9EAF7"
)

rejected_fill = PatternFill(
    fill_type="solid",
    start_color="F4CCCC",
    end_color="F4CCCC"
)

dashboard_sheet = workbook.active
dashboard_sheet.title = "Dashboard"

dashboard_sheet["A1"] = "EDI Processing Dashboard"
dashboard_sheet["A1"].font = title_font

dashboard_sheet["A3"] = "Total Transactions"
dashboard_sheet["B3"] = total_transactions

dashboard_sheet["A4"] = "Accepted"
dashboard_sheet["B4"] = accepted_count

dashboard_sheet["A5"] = "Rejected"
dashboard_sheet["B5"] = rejected_count

dashboard_sheet["A6"] = "Acceptance Rate"
dashboard_sheet["B6"] = acceptance_rate
dashboard_sheet["B6"].number_format = "0.0%"

dashboard_sheet["A8"] = "Files Processed"
dashboard_sheet["B8"] = len(file_status)

dashboard_sheet["A9"] = "Fully Accepted Files"
dashboard_sheet["B9"] = len(fully_accepted_files)

dashboard_sheet["A10"] = "Files with Rejections"
dashboard_sheet["B10"] = len(files_with_rejections)

dashboard_sheet["A13"] = "Rejection Reasons"
dashboard_sheet["A13"].font = section_font
dashboard_sheet["A13"].fill = header_fill

dashboard_sheet["A14"] = "Rejection Reason"
dashboard_sheet["B14"] = "Count"

dashboard_sheet["A14"].font = section_font
dashboard_sheet["B14"].font = section_font

dashboard_sheet["A14"].fill = header_fill
dashboard_sheet["B14"].fill = header_fill

dashboard_row = 15

for rejection_key, count in sorted(
    rejection_summary.items(),
    key=lambda item: item[1],
    reverse=True
):
    status_code = rejection_key[0]
    status_message = rejection_key[1]

    dashboard_sheet[f"A{dashboard_row}"] = (
        f"{status_code} - {status_message}"
    )
    dashboard_sheet[f"B{dashboard_row}"] = count

    dashboard_row += 1

for cell in ["A3", "A4", "A5", "A6", "A8", "A9", "A10"]:
    dashboard_sheet[cell].font = section_font
    dashboard_sheet[cell].fill = header_fill

for column in dashboard_sheet.columns:
    length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
    dashboard_sheet.column_dimensions[
        get_column_letter(column[0].column)
    ].width = length + 3

summary_sheet = workbook.create_sheet("Summary")

summary_sheet["A1"] = "EDI Daily Processing Report"
summary_sheet["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

summary_sheet["A1"].font = title_font

summary_sheet["A4"].font = section_font
summary_sheet["A10"].font = section_font

summary_sheet["A4"].fill = header_fill
summary_sheet["A10"].fill = header_fill

summary_sheet["A4"] = "Processing Summary"
summary_sheet["A5"] = "Total Transactions"
summary_sheet["B5"] = total_transactions

summary_sheet["A6"] = "Accepted"
summary_sheet["B6"] = accepted_count

summary_sheet["A7"] = "Rejected"
summary_sheet["B7"] = rejected_count

summary_sheet["A8"] = "Acceptance Rate"
summary_sheet["B8"] = f"{acceptance_rate:.1%}"

summary_sheet["A10"] = "File Summary"
summary_sheet["A11"] = "Files Processed"
summary_sheet["B11"] = len(file_status)

summary_sheet["A12"] = "Fully Accepted Files"
summary_sheet["B12"] = len(fully_accepted_files)

summary_sheet["A13"] = "Files with Rejections"
summary_sheet["B13"] = len(files_with_rejections)

summary_sheet.freeze_panes = "A4"

for column in summary_sheet.columns:
    length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
    summary_sheet.column_dimensions[
        get_column_letter(column[0].column)
    ].width = length + 3

raw_sheet = workbook.create_sheet("Raw Transactions")

with open(input_file, "r") as file:
    reader = csv.reader(file)

    for row in reader:
        raw_sheet.append(row)

for cell in raw_sheet[1]:
    cell.font = section_font
    cell.fill = header_fill

for row in raw_sheet.iter_rows(min_row=2):
    status_cell = row[8]

    if status_cell.value == "Rejected":
        for cell in row:
            cell.fill = rejected_fill

raw_sheet.freeze_panes = "A2"
raw_sheet.auto_filter.ref = raw_sheet.dimensions

for column in raw_sheet.columns:
    length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
    raw_sheet.column_dimensions[
        get_column_letter(column[0].column)
    ].width = length + 3

rejection_sheet = workbook.create_sheet("Rejection Details")

rejection_sheet.append([
    "file_name",
    "status_code",
    "status_message",
    "rejection_count"
])

for cell in rejection_sheet[1]:
    cell.font = section_font
    cell.fill = header_fill

for file_name in sorted(rejection_detail_by_file):
    for rejection_key, count in sorted(rejection_detail_by_file[file_name].items()):
        status_code = rejection_key[0]
        status_message = rejection_key[1]

        rejection_sheet.append([
            file_name,
            status_code,
            status_message,
            count
        ])

rejection_sheet.freeze_panes = "A2"

for column in rejection_sheet.columns:
    length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
    rejection_sheet.column_dimensions[
        get_column_letter(column[0].column)
    ].width = length + 3

chart = BarChart()
chart.title = "Accepted vs Rejected"
chart.y_axis.title = "Transaction Count"
chart.x_axis.title = "Status"

data = Reference(dashboard_sheet, min_col=2, min_row=4, max_row=5)
categories = Reference(dashboard_sheet, min_col=1, min_row=4, max_row=5)

chart.add_data(data)
chart.set_categories(categories)

dashboard_sheet.add_chart(chart, "D3")

rejection_chart = BarChart()
rejection_chart.type = "bar"
rejection_chart.style = 10
rejection_chart.title = "Top Rejection Reasons"
rejection_chart.x_axis.title = None
rejection_chart.y_axis.title = None

rejection_data = Reference(
    dashboard_sheet,
    min_col=2,
    min_row=14,
    max_row=dashboard_row - 1
)

rejection_categories = Reference(
    dashboard_sheet,
    min_col=1,
    min_row=15,
    max_row=dashboard_row - 1
)

rejection_chart.add_data(
    rejection_data,
    titles_from_data=True
)

rejection_chart.set_categories(rejection_categories)

rejection_chart.series[0].graphicalProperties.solidFill = "5B9BD5"

rejection_chart.legend = None

rejection_chart.width = 14
rejection_chart.height = 7

dashboard_sheet.add_chart(rejection_chart, "D16")

workbook.save(output_file)

print(f"Excel report created: {output_file}")